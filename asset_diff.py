#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
asset_diff.py — 资产明细差异分析脚本

功能：接收两个不同时期的资产明细Excel文件，逐条比较资产数据，
      计算各维度汇总差异，输出结构化差异报告。

用法：
    python asset_diff.py --current <当前期文件> --previous <对比期文件> --output <输出文件>

示例：
    python asset_diff.py --current "2026Q1.xlsx" --previous "2025Q4.xlsx" --output "差异报告.xlsx"
"""

import argparse
import time
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.utils import get_column_letter


# ============================================================
# 一、常量定义
# ============================================================

# 使用状态分类映射（大类，细分类）
STATUS_CATEGORY_MAP = {
    "①自用-自用办公": ("可开发运作", "自用办公"),
    "①自用-自用配套": ("可开发运作", "自用配套"),
    "②自营-独立运营": ("可开发运作", "独立运营"),
    "②自营-合作运营": ("可开发运作", "合作运营"),
    "③借用": ("暂不可开发利用", "借用"),
    "④闲置-危旧封闭": ("暂不可开发利用", "危旧封闭"),
    "④闲置-拟安置（已明确用于安置）": ("暂不可开发利用", "拟安置"),
    "⑤已出租-协议出租": ("可开发运作", "协议出租"),
    "⑤已出租-公开出租": ("可开发运作", "公开出租"),
    "⑥项目改造-待改造": ("可开发运作", "待改造"),
    "⑥项目改造-改造中": ("可开发运作", "改造中"),
    "⑦公共配套": ("暂不可开发利用", "公共配套"),
    "⑧房改预售": ("暂不可开发利用", "房改预售"),
    "⑨其他-未实际接收接管": ("暂不可开发利用", "未实际接收"),
    "⑨其他-待解决纠纷/产权不清": ("暂不可开发利用", "待解决纠纷/产权不清"),
    "⑨其他-征迁中": ("暂不可开发利用", "征迁中"),
    "⑩空置-已挂牌": ("可开发运作", "已挂牌"),
    "⑩空置-待挂牌出租": ("可开发运作", "待挂牌出租"),
    "⑩空置-待协议出租": ("可开发运作", "待协议出租"),
    "⑪职工宿舍（包括原来厂区职工宿舍）": ("暂不可开发利用", "职工宿舍"),
}

# 兼容旧版状态值（不含子分类）
STATUS_CATEGORY_MAP_COMPAT = {
    "①自用": ("可开发运作", "自用办公"),
    "②自营": ("可开发运作", "独立运营"),
    "③借用": ("暂不可开发利用", "借用"),
    "④闲置": ("暂不可开发利用", "危旧封闭"),
    "⑤已出租": ("可开发运作", "协议出租"),
    "⑥项目改造": ("可开发运作", "待改造"),
    "⑦公共配套": ("暂不可开发利用", "公共配套"),
    "⑧房改预售": ("暂不可开发利用", "房改预售"),
    "⑨其他": ("暂不可开发利用", "未实际接收"),
    "⑩空置": ("可开发运作", "已挂牌"),
    "⑪职工宿舍": ("暂不可开发利用", "职工宿舍"),
}

# 固定的20个合法资产状态类目（两期共有的标准状态值）
VALID_STATUS_VALUES = set(STATUS_CATEGORY_MAP.keys())

# 建筑类型标准化映射
BUILDING_TYPE_MAP = {
    "①商务(含写字楼、办公楼、酒店）": "商务",
    "①商务": "商务",
    "②公共配套": "公共配套",
    "③停车": "停车",
    "④住宅": "住宅",
    "⑤商场综合体": "商场",
    "⑥厂房": "厂房",
    "⑦仓库": "仓库",
    "⑧储藏间": "储藏间",
    "⑨店面": "店面",
    "⑩夹层": "夹层",
    "⑪文体": "文体",
    "⑫纯土地(空地)": "纯土地",
    "⑫纯土地": "纯土地",
    "⑬其他公共部分": "其它公共部分",
}

# 建筑类型排序（按经济分析表固定顺序）
BUILDING_TYPE_ORDER = [
    "商务", "公共配套", "停车", "住宅", "商场", "厂房",
    "仓库", "储藏间", "店面", "夹层", "文体", "纯土地", "其它公共部分",
]

# 暂不可开发利用的细分类列表
UNDEVELOPABLE_CATS = [
    "房改预售", "公共配套", "职工宿舍", "借用",
    "危旧封闭", "拟安置", "未实际接收", "待解决纠纷/产权不清", "征迁中",
]

# 可开发运作的细分类列表
DEVELOPABLE_CATS = [
    "自用办公", "自用配套", "独立运营", "合作运营",
    "协议出租", "公开出租", "已挂牌", "待挂牌出租", "待协议出租",
    "待改造", "改造中",
]

# 全部20个合法细分类（固定顺序，用于变动汇总等需要完整列出的场景）
ALL_VALID_CATEGORIES = UNDEVELOPABLE_CATS + DEVELOPABLE_CATS

# 已出租细分类
RENTED_CATS = ["协议出租", "公开出租"]

# 空置细分类
VACANT_CATS = ["已挂牌", "待挂牌出租", "待协议出租"]

# 自用（营）细分类
SELF_USE_CATS = ["自用办公", "自用配套", "独立运营", "合作运营"]

# 项目改造细分类
PROJECT_CATS = ["待改造", "改造中"]

# 无经营决策权8类
NO_DECISION_CATS = [
    "未实际接收", "房改预售", "公共配套", "待解决纠纷/产权不清",
    "拟安置", "借用", "职工宿舍", "征迁中",
]

# 数值型字段（读取时自动转换）
NUMERIC_FIELDS = [
    "建筑面积", "占地面积", "原始价值", "评估价值",
    "合同租金（月）", "合同租金（年）",
]


# ============================================================
# 二、数据读取与标准化
# ============================================================

def find_header_row(ws):
    """自动定位含'建筑面积'的行作为表头行"""
    for row in ws.iter_rows(min_row=1, max_row=10, max_col=ws.max_column):
        for cell in row:
            if cell.value and "建筑面积" in str(cell.value):
                return cell.row
    raise ValueError("未找到含'建筑面积'的表头行")


def build_column_names(ws, header_row):
    """
    合并父/子表头生成列名。
    策略：
    1. 子表头中如果含换行符，取换行前的部分作为列名（换行后通常是说明）
    2. 若子表头为空或是纯说明文字，则使用父表头的值
    """
    parent_row = header_row - 1
    col_names = []
    for col_idx in range(1, ws.max_column + 1):
        child_val = ws.cell(row=header_row, column=col_idx).value
        parent_val = ws.cell(row=parent_row, column=col_idx).value

        child_str = str(child_val).strip() if child_val else ""
        parent_str = str(parent_val).strip() if parent_val else ""

        # 处理换行：取换行前的部分作为列名
        if "\n" in child_str:
            child_str = child_str.split("\n")[0].strip()

        # 说明文字特征（整个值都是说明，不是列名）
        is_description = (
            "/" in child_str and col_idx <= 3  # 编码说明
        ) or (
            child_str.startswith("（") and "）" in child_str and len(child_str) > 10  # 括号说明
        ) or (
            child_str.count("/") >= 3  # 枚举值说明
        )

        if child_str and not is_description:
            name = child_str
        elif parent_str:
            name = parent_str
        else:
            name = f"col_{col_idx}"

        # 清理列名中残留的换行符
        name = name.replace("\n", "").strip()
        col_names.append(name)
    return col_names


def find_main_sheet(wb):
    """找到主表Sheet（名称含'古城不动产'或'总体情况'）"""
    for ws in wb.worksheets:
        name = ws.title
        if "古城不动产" in name or "总体情况" in name:
            return ws
    # 回退：使用第一个Sheet
    return wb.worksheets[0]


def read_asset_file(filepath):
    """读取资产明细Excel文件，返回标准化的DataFrame和基准日期"""
    filepath = Path(filepath)
    print(f"  正在读取: {filepath.name}")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = find_main_sheet(wb)
    sheet_name = ws.title
    print(f"  Sheet名称: {sheet_name}")

    # 读取基准日期（A2单元格）
    base_date = ""
    cell_a2 = ws.cell(row=2, column=1).value
    if cell_a2:
        base_date = str(cell_a2).replace("填报基准日：", "").strip()
    print(f"  基准日期: {base_date}")

    # 定位表头行
    header_row = find_header_row(ws)
    print(f"  表头行: 第{header_row}行")

    # 构建列名
    col_names = build_column_names(ws, header_row)
    print(f"  列数: {len(col_names)}")

    # 读取数据行（表头下一行开始）
    data_start = header_row + 1
    data_rows = []
    for row in ws.iter_rows(min_row=data_start, max_row=ws.max_row,
                            max_col=len(col_names), values_only=True):
        # 跳过全空行
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        data_rows.append(list(row))

    wb.close()

    df = pd.DataFrame(data_rows, columns=col_names)

    # 标准化列名映射
    df = standardize_columns(df)

    # 过滤有效行
    df = df[df["资产编码"].notna() & df["资产编码"].str.startswith("GC")].copy()
    print(f"  有效资产记录: {len(df)} 条")

    # 数值列转换
    for field in NUMERIC_FIELDS:
        if field in df.columns:
            df[field] = pd.to_numeric(df[field], errors="coerce")

    return df, base_date


def standardize_columns(df):
    """标准化列名，兼容不同版本的文件"""
    col_map = {}
    for col in df.columns:
        col_clean = str(col).strip()
        if "资产编码" in col_clean and "说明" not in col_clean:
            col_map[col] = "资产编码"
        elif col_clean in ("序号",):
            col_map[col] = "序号"
        elif col_clean in ("原始顺序",):
            col_map[col] = "原始顺序"
        elif "现使用单位" in col_clean:
            col_map[col] = "现使用单位"
        elif col_clean == "资产名称":
            col_map[col] = "资产名称"
        elif "资产分层" in col_clean:
            col_map[col] = "资产分层"
        elif "资产类别" in col_clean:
            col_map[col] = "资产类别"
        elif "建筑类型" in col_clean:
            col_map[col] = "建筑类型"
        elif "建成年份" in col_clean:
            col_map[col] = "建成年份"
        elif "是否公房" in col_clean:
            col_map[col] = "是否公房"
        elif "是否独栋" in col_clean:
            col_map[col] = "是否独栋"
        elif "所属古城分区" in col_clean:
            col_map[col] = "所属古城分区"
        elif "原产权单位" in col_clean:
            col_map[col] = "原产权单位"
        elif "划转依据" in col_clean:
            col_map[col] = "划转依据"
        elif "移交时间" in col_clean:
            col_map[col] = "移交时间"
        elif "所属街道" in col_clean:
            col_map[col] = "所属街道"
        elif "资产地理位置" in col_clean:
            col_map[col] = "资产地理位置"
        elif "建筑面积" in col_clean:
            col_map[col] = "建筑面积"
        elif "占地面积" in col_clean:
            col_map[col] = "占地面积"
        elif "建筑结构" in col_clean and "建筑结构" not in col_map:
            col_map[col] = "建筑结构"
        elif "土地类型" in col_clean:
            col_map[col] = "土地类型"
        elif "使用类型" in col_clean:
            col_map[col] = "使用类型"
        elif "是否属于古城范围" in col_clean:
            col_map[col] = "是否属于古城范围"
        elif "原始价值" in col_clean:
            col_map[col] = "原始价值"
        elif col_clean == "评估价值":
            col_map[col] = "评估价值"
        elif "评估单价" in col_clean:
            col_map[col] = "评估单价"
        elif "评估总价" in col_clean:
            col_map[col] = "评估总价"
        elif "是否入固定资产" in col_clean:
            col_map[col] = "是否入固定资产明细账"
        elif "未进" in col_clean and "造册" in col_clean:
            col_map[col] = "未进明细账是否造册管理"
        elif "不动产证号" in col_clean:
            col_map[col] = "不动产证号"
        elif "房产证号" in col_clean:
            col_map[col] = "房产证号"
        elif "土地证号" in col_clean:
            col_map[col] = "土地证号"
        elif col_clean == "管理公司":
            col_map[col] = "管理公司"
        elif col_clean == "管理小组":
            col_map[col] = "管理小组"
        elif "新管理小组" in col_clean:
            col_map[col] = "新管理小组"
        elif "资产状态" in col_clean:
            col_map[col] = "资产状态"
        elif "借用" in col_clean and "单位" not in col_clean and "时间" not in col_clean:
            col_map[col] = "借用"
        elif "借用单位" in col_clean:
            col_map[col] = "借用单位"
        elif "借用时间" in col_clean:
            col_map[col] = "借用时间"
        elif "是否进产权" in col_clean or "是否进交易" in col_clean:
            col_map[col] = "是否进产权"
        elif "承租户" in col_clean:
            col_map[col] = "承租户"
        elif "房租评估价" in col_clean:
            col_map[col] = "房租评估价（月）"
        elif "是否订立合同" in col_clean:
            col_map[col] = "是否订立合同"
        elif "押金" in col_clean:
            col_map[col] = "押金"
        elif "合同租金" in col_clean and "年" in col_clean:
            col_map[col] = "合同租金（年）"
        elif "合同租金" in col_clean and "月" in col_clean:
            col_map[col] = "合同租金（月）"
        elif "合同租金" in col_clean:
            # 默认为月租金
            col_map[col] = "合同租金（月）"
        elif "合同租赁起始" in col_clean:
            col_map[col] = "合同租赁起始"
        elif "合同租赁到期" in col_clean:
            col_map[col] = "合同租赁到期"
        elif "经营业态" in col_clean:
            col_map[col] = "经营业态"
        elif "租赁到期" in col_clean and "重新招租" in col_clean:
            col_map[col] = "租赁到期是否重新招租"
        elif "欠缴" in col_clean:
            col_map[col] = "累计欠缴租金"
        elif col_clean == "备注1":
            col_map[col] = "备注1"
        elif col_clean == "备注2":
            col_map[col] = "备注2"

    df = df.rename(columns=col_map)
    return df


def validate_data(df_current_raw, df_previous_raw, date_current, date_previous):
    """
    数据检验函数：检查两个时期的数据质量，输出检验报告。
    可独立修改和扩展检验规则。

    检验项：
    1. 资产状态类目检验：检查是否有不在固定20个合法类目中的状态值
    2. 资产编码唯一性检验
    3. 建筑面积缺失/异常检验
    4. 两期共有资产编码统计

    Returns:
        warnings (list): 警告信息列表
    """
    warnings = []
    print("  --- 数据检验 ---")

    # 检验1：资产状态类目
    for label, df, date in [("当前期", df_current_raw, date_current),
                             ("对比期", df_previous_raw, date_previous)]:
        statuses = df["资产状态"].dropna().unique()
        invalid = []
        for s in statuses:
            s_norm = _normalize_status(str(s).strip())
            matched = any(
                s in VALID_STATUS_VALUES or
                s_norm == _normalize_status(k)
                for k in VALID_STATUS_VALUES
            )
            if not matched:
                cnt = (df["资产状态"] == s).sum()
                invalid.append((s, cnt))
        if invalid:
            detail = ", ".join(f"{s}({cnt}条)" for s, cnt in invalid)
            msg = f"[警告] {label}({date}) 存在非标准状态类目: {detail}"
            warnings.append(msg)
            print(f"  {msg}")
            for s, cnt in invalid:
                print(f'    → "{s}"({cnt}条) 将映射为未知，不参与任何统计和比较')
        else:
            print(f"  {label}({date}) 资产状态类目: 全部 {len(statuses)} 种均在合法范围内 ✓")

    # 检验2：资产编码唯一性
    for label, df in [("当前期", df_current_raw), ("对比期", df_previous_raw)]:
        total = len(df)
        unique = df["资产编码"].nunique()
        if total != unique:
            msg = f"[警告] {label} 资产编码存在重复: 总{total}条, 去重{unique}条, 重复{total-unique}条"
            warnings.append(msg)
            print(f"  {msg}")
            print(f"    → 重复记录可能导致汇总统计偏大，建议检查数据源")
        else:
            print(f"  {label} 资产编码唯一性: {unique} 条无重复 ✓")

    # 检验3：建筑面积缺失/异常（包含0）
    for label, df in [("当前期", df_current_raw), ("对比期", df_previous_raw)]:
        missing = df["建筑面积"].isna().sum()
        zero = (df["建筑面积"] == 0).sum()
        negative = (df["建筑面积"] < 0).sum()
        problem = missing + zero + negative
        if problem > 0:
            parts = []
            if missing > 0:
                parts.append(f"缺失{missing}条")
            if zero > 0:
                parts.append(f"为0: {zero}条")
            if negative > 0:
                parts.append(f"为负: {negative}条")
            msg = f"[提示] {label} 建筑面积异常: {problem}条（{', '.join(parts)}，均按0处理）"
            warnings.append(msg)
            print(f"  {msg}")
            # 打印编码
            problem_df = df[df["建筑面积"].isna() | (df["建筑面积"] == 0) | (df["建筑面积"] < 0)]
            codes = problem_df["资产编码"].tolist()
            print(f"    → 编码: {', '.join(codes)}")
            print(f"    → 注：对于合并单元格的处理，仅第一条有建筑面积，其余为缺失")
        else:
            print(f"  {label} 建筑面积: 无缺失、无0、无负值 ✓")

    # 检验4：两期共有资产编码
    codes_c = set(df_current_raw["资产编码"].dropna().unique())
    codes_p = set(df_previous_raw["资产编码"].dropna().unique())
    common = codes_c & codes_p
    only_c = codes_c - codes_p
    only_p = codes_p - codes_c
    print(f"  两期资产编码对比: 共有 {len(common)} 条, 仅当前期新增 {len(only_c)} 条, 仅对比期删除 {len(only_p)} 条")

    if warnings:
        print(f"  --- 检验完成: {len(warnings)} 条警告，请关注上述提示 ---")
    else:
        print(f"  --- 检验完成: 全部通过 ✓ ---")

    return warnings


def split_self_owned(df):
    """拆分自有资产和非自有资产"""
    mask = df["原产权单位"].fillna("").str.contains("代政府征收|业主|征迁", na=False)
    df_self = df[~mask].copy()
    df_nonself = df[mask].copy()
    return df_self, df_nonself


def _normalize_status(status):
    """统一全角/半角括号，便于匹配"""
    return status.replace("）", ")").replace("（", "(")


def map_status_category(df):
    """映射资产状态为（大类，细分类）"""
    def _map(status):
        if pd.isna(status) or str(status).strip() == "":
            return ("未知", "未知")
        status = str(status).strip()
        status_norm = _normalize_status(status)
        # 先尝试精确匹配（原始值）
        if status in STATUS_CATEGORY_MAP:
            return STATUS_CATEGORY_MAP[status]
        # 再尝试标准化后的匹配
        for key, val in STATUS_CATEGORY_MAP.items():
            if status_norm == _normalize_status(key):
                return val
        # 兼容旧版状态值
        if status in STATUS_CATEGORY_MAP_COMPAT:
            return STATUS_CATEGORY_MAP_COMPAT[status]
        # 模糊匹配
        for key, val in STATUS_CATEGORY_MAP.items():
            if status_norm.startswith(_normalize_status(key[:2])):
                return val
        return ("未知", "未知")

    result = df["资产状态"].apply(_map)
    df["使用大类"] = result.apply(lambda x: x[0])
    df["使用细分类"] = result.apply(lambda x: x[1])
    return df


def standardize_building_type(df):
    """标准化建筑类型"""
    def _map(bt):
        if pd.isna(bt) or str(bt).strip() == "":
            return "未知"
        bt = str(bt).strip()
        if bt in BUILDING_TYPE_MAP:
            return BUILDING_TYPE_MAP[bt]
        # 模糊匹配
        for key, val in BUILDING_TYPE_MAP.items():
            if bt.startswith(key[:2]):
                return val
        return bt

    df["标准建筑类型"] = df["建筑类型"].apply(_map)
    return df


# ============================================================
# 三、统计指标计算
# ============================================================

def calc_summary_stats(df, label=""):
    """计算单期汇总统计指标（仅基于合法的20个状态类目）"""
    # 排除非标准状态类目的记录
    df = df[df["使用细分类"] != "未知"].copy()
    if len(df) == 0:
        return {}

    stats = {}
    total_area = df["建筑面积"].sum()
    total_count = df["资产编码"].nunique()

    stats["房产总面积"] = round(total_area, 2)
    stats["资产总宗数"] = total_count

    # 按使用细分类汇总
    cat_stats = df.groupby("使用细分类").agg(
        面积=("建筑面积", "sum"),
        宗数=("资产编码", "nunique"),
    ).fillna(0)
    cat_stats["面积"] = cat_stats["面积"].round(2)

    # 暂不可开发利用
    undevelopable = cat_stats.loc[
        cat_stats.index.isin(UNDEVELOPABLE_CATS)
    ]
    stats["暂不可开发利用面积"] = round(undevelopable["面积"].sum(), 2)
    stats["暂不可开发利用宗数"] = int(undevelopable["宗数"].sum())

    # 可开发运作
    developable = cat_stats.loc[
        cat_stats.index.isin(DEVELOPABLE_CATS)
    ]
    stats["可开发运作面积"] = round(developable["面积"].sum(), 2)
    stats["可开发运作宗数"] = int(developable["宗数"].sum())

    # 已出租
    rented = cat_stats.loc[cat_stats.index.isin(RENTED_CATS)]
    stats["已出租面积"] = round(rented["面积"].sum(), 2)
    stats["已出租宗数"] = int(rented["宗数"].sum())

    # 空置
    vacant = cat_stats.loc[cat_stats.index.isin(VACANT_CATS)]
    stats["空置面积"] = round(vacant["面积"].sum(), 2)
    stats["空置宗数"] = int(vacant["宗数"].sum())

    # 可出租
    stats["可出租面积"] = round(stats["已出租面积"] + stats["空置面积"], 2)

    # 出租率
    if stats["可出租面积"] > 0:
        stats["出租率"] = round(stats["已出租面积"] / stats["可出租面积"] * 100, 6)
    else:
        stats["出租率"] = 0.0

    # 在约合同数
    rented_df = df[df["使用细分类"].isin(RENTED_CATS)]
    contract_count = 0
    if "是否订立合同" in rented_df.columns:
        contract_count = int(
            rented_df["是否订立合同"].fillna("").str.strip().str.upper().eq("是").sum()
        )
    stats["在约合同数"] = contract_count

    # 自用（营）
    self_use = cat_stats.loc[cat_stats.index.isin(SELF_USE_CATS)]
    stats["自用（营）面积"] = round(self_use["面积"].sum(), 2)

    # 项目开发
    project = cat_stats.loc[cat_stats.index.isin(PROJECT_CATS)]
    stats["项目开发面积"] = round(project["面积"].sum(), 2)

    # 待改造 / 改造中
    pending = cat_stats.loc[cat_stats.index == "待改造"]
    ongoing = cat_stats.loc[cat_stats.index == "改造中"]
    stats["待改造面积"] = round(pending["面积"].sum(), 2) if len(pending) > 0 else 0.0
    stats["改造中面积"] = round(ongoing["面积"].sum(), 2) if len(ongoing) > 0 else 0.0

    # 月租金合计
    # stats["月租金合计"] = round(rented_df["合同租金（月）"].sum(), 2) if "合同租金（月）" in rented_df.columns else 0.0

    # 无经营决策权资产面积
    no_decision = cat_stats.loc[cat_stats.index.isin(NO_DECISION_CATS)]
    stats["无经营决策权面积"] = round(no_decision["面积"].sum(), 2)
    stats["可自主决策经营面积"] = round(total_area - stats["无经营决策权面积"], 2)

    # 剔除后面积（用于口径2）
    exclude_cats = ["未实际接收", "公共配套", "房改预售", "征迁中", "职工宿舍", "待解决纠纷/产权不清"]
    excluded = cat_stats.loc[cat_stats.index.isin(exclude_cats)]
    stats["剔除后面积"] = round(total_area - excluded["面积"].sum(), 2)

    # 口径1/2/3
    if total_area > 0:
        stats["出租率_口径1"] = round(stats["已出租面积"] / total_area * 100, 6)
    else:
        stats["出租率_口径1"] = 0.0
    if stats["剔除后面积"] > 0:
        stats["出租率_口径2"] = round(stats["已出租面积"] / stats["剔除后面积"] * 100, 6)
    else:
        stats["出租率_口径2"] = 0.0
    stats["出租率_口径3"] = stats["出租率"]  # 已出租/可出租

    # 按资产类别（I类/II类）
    asset_cat_stats = df.groupby("资产类别").agg(
        面积=("建筑面积", "sum"),
        宗数=("资产编码", "nunique"),
    ).fillna(0)
    asset_cat_stats["面积"] = asset_cat_stats["面积"].round(2)
    stats["I类面积"] = round(asset_cat_stats.loc["I类", "面积"], 2) if "I类" in asset_cat_stats.index else 0.0
    stats["II类面积"] = round(asset_cat_stats.loc["II类", "面积"], 2) if "II类" in asset_cat_stats.index else 0.0

    return stats


def _safe_area(val):
    """安全获取面积值，将NaN/None转为0"""
    v = val if val else 0
    return 0 if pd.isna(v) else v


# ============================================================
# 四、逐条资产比较
# ============================================================

def compare_assets_by_status(df_current, df_previous):
    """
    按使用状态分类比较两个时期的资产，返回变更明细DataFrame。
    变更类型：新增、删除、面积调整。
    """
    df_current = df_current[df_current["使用细分类"] != "未知"].copy()
    df_previous = df_previous[df_previous["使用细分类"] != "未知"].copy()

    current_codes = set(df_current["资产编码"].unique())
    previous_codes = set(df_previous["资产编码"].unique())

    added_codes = current_codes - previous_codes
    removed_codes = previous_codes - current_codes
    common_codes = current_codes & previous_codes

    changes = []

    # --- 新增资产 ---
    for code in sorted(added_codes):
        row = df_current[df_current["资产编码"] == code].iloc[0]
        area = _safe_area(row.get("建筑面积"))
        changes.append({
            "资产编码": code,
            "资产分层（同系统名称）": row.get("资产分层", "") or row.get("资产名称", ""),
            "管理小组": row.get("新管理小组", ""),
            "建筑类型": row.get("标准建筑类型", ""),
            "变更类型": "新增",
            "建筑面积（当前期）": area,
            "建筑面积（对比期）": 0,
            "建筑面积（变动）": area,
            "使用状态（当前期）": row.get("使用细分类", ""),
            "使用状态（对比期）": "",
            "变更详情": "新增资产",
        })

    # --- 删除资产 ---
    for code in sorted(removed_codes):
        row = df_previous[df_previous["资产编码"] == code].iloc[0]
        area = _safe_area(row.get("建筑面积"))
        changes.append({
            "资产编码": code,
            "资产分层（同系统名称）": row.get("资产分层", "") or row.get("资产名称", ""),
            "管理小组": row.get("新管理小组", ""),
            "建筑类型": row.get("标准建筑类型", ""),
            "变更类型": "删除",
            "建筑面积（当前期）": 0,
            "建筑面积（对比期）": area,
            "建筑面积（变动）": -area,
            "使用状态（当前期）": "",
            "使用状态（对比期）": row.get("使用细分类", ""),
            "变更详情": "删除资产",
        })

    # --- 面积调整资产（两期都有，状态或面积有变化） ---
    for code in sorted(common_codes):
        row_c = df_current[df_current["资产编码"] == code].iloc[0]
        row_p = df_previous[df_previous["资产编码"] == code].iloc[0]

        area_c = _safe_area(row_c.get("建筑面积"))
        area_p = _safe_area(row_p.get("建筑面积"))
        area_diff = round(area_c - area_p, 2)

        subcat_c = row_c.get("使用细分类", "")
        subcat_p = row_p.get("使用细分类", "")

        status_changed = subcat_c != subcat_p
        area_changed = abs(area_diff) > 0.01

        if status_changed or area_changed:
            detail_parts = []
            if status_changed:
                detail_parts.append(f"使用状态: {subcat_p}→{subcat_c}")
            if area_changed:
                detail_parts.append(f"面积: {area_p}→{area_c}㎡({area_diff:+.2f})")

            changes.append({
                "资产编码": code,
                "资产分层（同系统名称）": row_c.get("资产分层", "") or row_c.get("资产名称", ""),
                "管理小组": row_c.get("新管理小组", ""),
                "建筑类型": row_c.get("标准建筑类型", ""),
                "变更类型": "状态变更" if status_changed else "面积调整",
                "建筑面积（当前期）": area_c,
                "建筑面积（对比期）": area_p,
                "建筑面积（变动）": area_diff,
                "使用状态（当前期）": subcat_c,
                "使用状态（对比期）": subcat_p,
                "变更详情": "；".join(detail_parts),
            })

    return pd.DataFrame(changes)


def compare_assets_by_building_type(df_current, df_previous):
    """
    按建筑类型比较两个时期的资产，返回变更明细DataFrame。
    变更类型：新增、删除、面积调整。
    """
    df_current = df_current[df_current["使用细分类"] != "未知"].copy()
    df_previous = df_previous[df_previous["使用细分类"] != "未知"].copy()

    current_codes = set(df_current["资产编码"].unique())
    previous_codes = set(df_previous["资产编码"].unique())

    added_codes = current_codes - previous_codes
    removed_codes = previous_codes - current_codes
    common_codes = current_codes & previous_codes

    changes = []

    # --- 新增资产 ---
    for code in sorted(added_codes):
        row = df_current[df_current["资产编码"] == code].iloc[0]
        area = _safe_area(row.get("建筑面积"))
        changes.append({
            "资产编码": code,
            "资产分层（同系统名称）": row.get("资产分层", "") or row.get("资产名称", ""),
            "管理小组": row.get("新管理小组", ""),
            "使用状态": row.get("使用细分类", ""),
            "变更类型": "新增",
            "建筑面积（当前期）": area,
            "建筑面积（对比期）": 0,
            "建筑面积（变动）": area,
            "建筑类型（当前期）": row.get("标准建筑类型", ""),
            "建筑类型（对比期）": "",
            "变更详情": "新增资产",
        })

    # --- 删除资产 ---
    for code in sorted(removed_codes):
        row = df_previous[df_previous["资产编码"] == code].iloc[0]
        area = _safe_area(row.get("建筑面积"))
        changes.append({
            "资产编码": code,
            "资产分层（同系统名称）": row.get("资产分层", "") or row.get("资产名称", ""),
            "管理小组": row.get("新管理小组", ""),
            "使用状态": row.get("使用细分类", ""),
            "变更类型": "删除",
            "建筑面积（当前期）": 0,
            "建筑面积（对比期）": area,
            "建筑面积（变动）": -area,
            "建筑类型（当前期）": "",
            "建筑类型（对比期）": row.get("标准建筑类型", ""),
            "变更详情": "删除资产",
        })

    # --- 面积调整资产（两期都有，类型或面积有变化） ---
    for code in sorted(common_codes):
        row_c = df_current[df_current["资产编码"] == code].iloc[0]
        row_p = df_previous[df_previous["资产编码"] == code].iloc[0]

        area_c = _safe_area(row_c.get("建筑面积"))
        area_p = _safe_area(row_p.get("建筑面积"))
        area_diff = round(area_c - area_p, 2)

        bt_c = row_c.get("标准建筑类型", "")
        bt_p = row_p.get("标准建筑类型", "")

        type_changed = bt_c != bt_p
        area_changed = abs(area_diff) > 0.01

        if type_changed or area_changed:
            detail_parts = []
            if type_changed:
                detail_parts.append(f"建筑类型: {bt_p}→{bt_c}")
            if area_changed:
                detail_parts.append(f"面积: {area_p}→{area_c}㎡({area_diff:+.2f})")

            changes.append({
                "资产编码": code,
                "资产分层（同系统名称）": row_c.get("资产分层", "") or row_c.get("资产名称", ""),
                "管理小组": row_c.get("新管理小组", ""),
                "使用状态": row_c.get("使用细分类", ""),
                "变更类型": "类型变更" if type_changed else "面积调整",
                "建筑面积（当前期）": area_c,
                "建筑面积（对比期）": area_p,
                "建筑面积（变动）": area_diff,
                "建筑类型（当前期）": bt_c,
                "建筑类型（对比期）": bt_p,
                "变更详情": "；".join(detail_parts),
            })

    return pd.DataFrame(changes)


def _write_detail_sheet(ws, headers, changes_df, title):
    """写入变动明细表（Sheet 4/5 公共逻辑）"""
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    # 表头
    header_row = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=i, value=h)
    _style_header(ws, header_row, len(headers))

    # 文本列名集合（用于对齐判断）
    text_cols = {h for h in headers if not any(k in h for k in ("面积",))}

    # 数据行
    data_start = header_row + 1
    row = data_start
    for _, r in changes_df.iterrows():
        for i, h in enumerate(headers, 1):
            val = r.get(h, "")
            ws.cell(row=row, column=i, value=val).border = _THIN_BORDER
            if h in text_cols:
                ws.cell(row=row, column=i).alignment = _TEXT_ALIGN
            else:
                ws.cell(row=row, column=i).alignment = _NUM_ALIGN
                ws.cell(row=row, column=i).number_format = '0.00'

        change_type = r.get("变更类型", "")
        if change_type == "新增":
            ws.cell(row=row, column=5).fill = _INCREASE_FILL
        elif change_type == "删除":
            ws.cell(row=row, column=5).fill = _DECREASE_FILL
        elif change_type in ("面积调整", "状态变更", "类型变更"):
            ws.cell(row=row, column=5).fill = _HIGHLIGHT_FILL
        row += 1

    # 汇总行
    data_end = row - 1
    ws.cell(row=row, column=1, value="合计").font = Font(bold=True, size=11)
    ws.cell(row=row, column=1).alignment = _TEXT_ALIGN
    for i in range(1, len(headers) + 1):
        ws.cell(row=row, column=i).border = _THIN_BORDER
        ws.cell(row=row, column=i).font = Font(bold=True, size=11)
    for i, h in enumerate(headers, 1):
        if h in text_cols:
            ws.cell(row=row, column=i).alignment = _TEXT_ALIGN
            continue
        col_letter = get_column_letter(i)
        ws.cell(row=row, column=i).value = f"=SUM({col_letter}{data_start}:{col_letter}{data_end})"
        ws.cell(row=row, column=i).alignment = _NUM_ALIGN
        if "面积" in h:
            ws.cell(row=row, column=i).number_format = '0.00'

    # 冻结标题行
    ws.freeze_panes = f"A{data_start}"
    _auto_width(ws, len(headers), row)


# ============================================================
# 五、差异归因分析
# ============================================================

def analyze_category_changes(df_current, df_previous, category_col="使用细分类"):
    """
    分析各分类的面积变动，直接从原始数据计算，不依赖变更明细。
    对每个分类，将资产分为三组：
      - 新增：当前期属于该分类，对比期不属于
      - 删除：对比期属于该分类，当前期不属于
      - 面积调整：两期都属于该分类，但面积有变化
    """
    results = []

    # 排除未知
    df_current = df_current[df_current[category_col] != "未知"].copy()
    df_previous = df_previous[df_previous[category_col] != "未知"].copy()

    # 使用固定的20个标准类目
    all_cats = list(ALL_VALID_CATEGORIES)

    for cat in all_cats:
        # 该分类在两个时期的资产
        curr_codes = set(df_current[df_current[category_col] == cat]["资产编码"].unique())
        prev_codes = set(df_previous[df_previous[category_col] == cat]["资产编码"].unique())

        # 面积和宗数（缺失面积视为0）
        area_current = round(df_current[df_current[category_col] == cat]["建筑面积"].fillna(0).sum(), 2) if curr_codes else 0.0
        area_previous = round(df_previous[df_previous[category_col] == cat]["建筑面积"].fillna(0).sum(), 2) if prev_codes else 0.0
        count_current = len(curr_codes)
        count_previous = len(prev_codes)
        diff = round(area_current - area_previous, 2)

        # 新增：当前期属于该分类，对比期不属于
        added_codes = curr_codes - prev_codes
        新增面积 = round(df_current[df_current["资产编码"].isin(added_codes) & (df_current[category_col] == cat)]["建筑面积"].fillna(0).sum(), 2)

        # 删除：对比期属于该分类，当前期不属于
        removed_codes = prev_codes - curr_codes
        删除面积 = round(df_previous[df_previous["资产编码"].isin(removed_codes) & (df_previous[category_col] == cat)]["建筑面积"].fillna(0).sum(), 2)

        # 面积调整：两期都属于该分类，但面积有变化（缺失面积视为0）
        common_codes = curr_codes & prev_codes
        if common_codes:
            common_df = df_current[df_current["资产编码"].isin(common_codes) & (df_current[category_col] == cat)].merge(
                df_previous[df_previous["资产编码"].isin(common_codes) & (df_previous[category_col] == cat)],
                on="资产编码", suffixes=("_c", "_p")
            )
            common_df["建筑面积_c"] = common_df["建筑面积_c"].fillna(0)
            common_df["建筑面积_p"] = common_df["建筑面积_p"].fillna(0)
            common_df["面积差"] = common_df["建筑面积_c"] - common_df["建筑面积_p"]
            area_changed = common_df[common_df["面积差"].abs() > 0.01]
            面积调整宗数 = len(area_changed)
            面积调整额 = round(area_changed["面积差"].sum(), 2)
        else:
            面积调整宗数 = 0
            面积调整额 = 0.0

        results.append({
            "使用状态": cat,
            "当前期宗数": count_current,
            "对比期宗数": count_previous,
            "当前期面积": area_current,
            "对比期面积": area_previous,
            "变动面积": diff,
            "新增宗数": len(added_codes),
            "新增面积": 新增面积,
            "删除宗数": len(removed_codes),
            "删除面积": 删除面积,
            "面积调整宗数": 面积调整宗数,
            "面积调整额": 面积调整额,
        })

    return pd.DataFrame(results)




def calc_building_type_summary(df_current, df_previous):
    """计算建筑类型变动汇总，返回字典列表。

    每个字典包含: name, count_c, count_p, area_c, area_p, area_diff,
    added_count, added_area, removed_count, removed_area, adj_count, adj_area
    """
    df_curr_bt = df_current[df_current["使用细分类"] != "未知"].copy()
    df_prev_bt = df_previous[df_previous["使用细分类"] != "未知"].copy()

    all_types = list(BUILDING_TYPE_ORDER) + sorted(
        (set(df_curr_bt["标准建筑类型"].unique()) | set(df_prev_bt["标准建筑类型"].unique())) - set(BUILDING_TYPE_ORDER)
    )

    result = []
    for bt in all_types:
        if bt == "未知":
            continue
        curr_codes = set(df_curr_bt[df_curr_bt["标准建筑类型"] == bt]["资产编码"].unique())
        prev_codes = set(df_prev_bt[df_prev_bt["标准建筑类型"] == bt]["资产编码"].unique())

        area_current = round(df_curr_bt[df_curr_bt["标准建筑类型"] == bt]["建筑面积"].fillna(0).sum(), 2) if curr_codes else 0.0
        area_previous = round(df_prev_bt[df_prev_bt["标准建筑类型"] == bt]["建筑面积"].fillna(0).sum(), 2) if prev_codes else 0.0
        count_current = len(curr_codes)
        count_previous = len(prev_codes)
        diff = round(area_current - area_previous, 2)

        added_codes = curr_codes - prev_codes
        added_area = round(df_curr_bt[df_curr_bt["资产编码"].isin(added_codes) & (df_curr_bt["标准建筑类型"] == bt)]["建筑面积"].fillna(0).sum(), 2)

        removed_codes = prev_codes - curr_codes
        removed_area = round(df_prev_bt[df_prev_bt["资产编码"].isin(removed_codes) & (df_prev_bt["标准建筑类型"] == bt)]["建筑面积"].fillna(0).sum(), 2)

        common_codes = curr_codes & prev_codes
        if common_codes:
            common_df = df_curr_bt[df_curr_bt["资产编码"].isin(common_codes) & (df_curr_bt["标准建筑类型"] == bt)].merge(
                df_prev_bt[df_prev_bt["资产编码"].isin(common_codes) & (df_prev_bt["标准建筑类型"] == bt)],
                on="资产编码", suffixes=("_c", "_p")
            )
            common_df["建筑面积_c"] = common_df["建筑面积_c"].fillna(0)
            common_df["建筑面积_p"] = common_df["建筑面积_p"].fillna(0)
            common_df["面积差"] = common_df["建筑面积_c"] - common_df["建筑面积_p"]
            area_changed = common_df[common_df["面积差"].abs() > 0.01]
            adj_count = len(area_changed)
            adj_area = round(area_changed["面积差"].sum(), 2)
        else:
            adj_count = 0
            adj_area = 0.0

        result.append({
            "name": bt, "count_c": count_current, "count_p": count_previous,
            "area_c": area_current, "area_p": area_previous, "area_diff": diff,
            "added_count": len(added_codes), "added_area": added_area,
            "removed_count": len(removed_codes), "removed_area": removed_area,
            "adj_count": adj_count, "adj_area": adj_area,
        })
    return result


# ============================================================
# 六、Excel 样式常量与辅助函数
# ============================================================

_HEADER_FONT = Font(bold=True, size=11)
_HEADER_FILL = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_NUM_ALIGN = Alignment(horizontal="right", vertical="center")
_TEXT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
_THIN_BORDER = Border(
    left=openpyxl.styles.Side(style="thin"),
    right=openpyxl.styles.Side(style="thin"),
    top=openpyxl.styles.Side(style="thin"),
    bottom=openpyxl.styles.Side(style="thin"),
)
_HIGHLIGHT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
_INCREASE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_DECREASE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def _style_header(ws, row, max_col):
    """设置表头行样式并添加筛选"""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER
    ws.auto_filter.ref = f"A{row}:{get_column_letter(max_col)}{row}"


def _auto_width(ws, max_col, max_row, min_width=10, max_width=50):
    """自动调整列宽"""
    for col in range(1, max_col + 1):
        max_len = min_width
        for row in range(1, min(max_row + 1, 100)):
            val = ws.cell(row=row, column=col).value
            if val:
                length = sum(2 if ord(c) > 127 else 1 for c in str(val))
                max_len = max(max_len, min(length + 2, max_width))
        ws.column_dimensions[get_column_letter(col)].width = max_len


# ============================================================
# 七、输出报告
# ============================================================

def write_excel_report(output_path, stats_current, stats_previous,
                       changes_status_df, changes_bt_df, cat_changes_df,
                       date_current, date_previous,
                       df_current, df_previous,
                       force=False):
    """输出差异报告Excel"""

    wb = openpyxl.Workbook()

    # ============================================================
    # Sheet 1: 汇总对比
    # ============================================================
    ws1 = wb.active
    ws1.title = "汇总对比"

    # 标题
    ws1.merge_cells("A1:F1")
    ws1.cell(row=1, column=1, value="资产统计指标汇总对比").font = Font(bold=True, size=14)
    ws1.cell(row=2, column=1, value=f"当前期: {date_current}")
    ws1.cell(row=3, column=1, value=f"对比期: {date_previous}")

    # 表头
    headers = ["指标", "当前期", "对比期", "变动值", "变动率(%)", "说明"]
    for i, h in enumerate(headers, 1):
        ws1.cell(row=5, column=i, value=h)
    _style_header(ws1, 5, len(headers))

    # 指标定义：(key, label, unit, is_rate, description)
    indicators = [
        ("房产总面积", "房产总面积", "㎡", False, "全部自有资产建筑面积之和"),
        ("资产总宗数", "资产总宗数", "宗", False, "资产编码去重计数"),
        ("暂不可开发利用面积", "暂不可开发利用面积", "㎡", False, "房改预售、公共配套、职工宿舍、借用、危旧封闭、拟安置、未实际接收、待解决纠纷/产权不清、征迁中"),
        ("可开发运作面积", "可开发运作面积", "㎡", False, "自用办公、自用配套、独立运营、合作运营、协议出租、公开出租、已挂牌、待挂牌出租、待协议出租、待改造、改造中"),
        ("已出租面积", "已出租面积", "㎡", False, "协议出租+公开出租"),
        ("已出租宗数", "已出租宗数", "宗", False, ""),
        ("空置面积", "空置面积", "㎡", False, "已挂牌+待挂牌出租+待协议出租"),
        ("空置宗数", "空置宗数", "宗", False, ""),
        ("可出租面积", "可出租面积", "㎡", False, "已出租面积+空置面积"),
        ("出租率", "出租率(口径3)", "%", True, "已出租面积/可出租面积"),
        ("在约合同数", "在约合同数", "份", False, "已出租中是否订立合同=是的行数"),
        ("自用（营）面积", "自用（营）面积", "㎡", False, "自用办公+自用配套+独立运营+合作运营"),
        ("项目开发面积", "项目开发面积", "㎡", False, "待改造+改造中"),
        ("待改造面积", "待改造面积", "㎡", False, ""),
        ("改造中面积", "改造中面积", "㎡", False, ""),
        # ("月租金合计", "月租金合计", "元", False, "已出租资产合同租金月之和"),
        ("可自主决策经营面积", "可自主决策经营面积", "㎡", False, "自用办公、自用配套、独立运营、合作运营、协议出租、公开出租、已挂牌、待挂牌出租、待协议出租、待改造、改造中、危旧封闭"),
        ("I类面积", "I类面积", "㎡", False, "资产类别=I类"),
        ("II类面积", "II类面积", "㎡", False, "资产类别=II类"),
        ("无经营决策权面积", "无经营决策权面积", "㎡", False, "未实际接收、房改预售、公共配套、待解决纠纷/产权不清、拟安置、借用、职工宿舍、征迁中"),
    ]

    row = 6
    for key, label, unit, is_rate, desc in indicators:
        val_c = stats_current.get(key, 0)
        val_p = stats_previous.get(key, 0)
        diff = round(val_c - val_p, 2) if not is_rate else round(val_c - val_p, 6)

        ws1.cell(row=row, column=1, value=label).alignment = _TEXT_ALIGN
        ws1.cell(row=row, column=2, value=val_c).alignment = _NUM_ALIGN
        ws1.cell(row=row, column=3, value=val_p).alignment = _NUM_ALIGN

        if is_rate:
            # 百分比指标：值存为小数（如0.9363），B/C列显示为百分比，D/E列为百分点变动
            ws1.cell(row=row, column=2, value=val_c / 100).alignment = _NUM_ALIGN
            ws1.cell(row=row, column=2).number_format = '0.00%'
            ws1.cell(row=row, column=3, value=val_p / 100).alignment = _NUM_ALIGN
            ws1.cell(row=row, column=3).number_format = '0.00%'
            ws1.cell(row=row, column=4).value = f"=(B{row}-C{row})*100"
            ws1.cell(row=row, column=4).number_format = '0.00"%"'
            ws1.cell(row=row, column=5).value = f"=D{row}"
            ws1.cell(row=row, column=5).number_format = '0.00"%"'
        else:
            ws1.cell(row=row, column=4).value = f"=B{row}-C{row}"
            ws1.cell(row=row, column=4).number_format = '0.00'
            ws1.cell(row=row, column=5).value = f'=IF(C{row}=0,"-",D{row}/ABS(C{row}))'
            ws1.cell(row=row, column=5).number_format = '0.00%'

        ws1.cell(row=row, column=4).alignment = _NUM_ALIGN
        ws1.cell(row=row, column=5).alignment = _NUM_ALIGN
        ws1.cell(row=row, column=6, value=desc).alignment = _TEXT_ALIGN

        # 高亮变动
        if abs(diff) > 0.01:
            fill = _INCREASE_FILL if diff > 0 else _DECREASE_FILL
            for c in range(2, 6):
                ws1.cell(row=row, column=c).fill = fill

        for c in range(1, 7):
            ws1.cell(row=row, column=c).border = _THIN_BORDER

        row += 1

    _auto_width(ws1, len(headers), row)

    # ============================================================
    # Sheet 2: 分类变动汇总
    # ============================================================
    ws2 = wb.create_sheet("使用状态变动汇总")

    ws2.merge_cells("A1:L1")
    ws2.cell(row=1, column=1, value="按使用状态分类变动汇总").font = Font(bold=True, size=14)

    headers2 = [
        "使用状态", "当前期宗数", "对比期宗数", "当前期面积", "对比期面积", "变动面积",
        "新增宗数", "新增面积", "删除宗数", "删除面积", "面积调整宗数", "面积调整额",
    ]
    for i, h in enumerate(headers2, 1):
        ws2.cell(row=3, column=i, value=h)
    _style_header(ws2, 3, len(headers2))

    row = 4
    for _, r in cat_changes_df.iterrows():
        for i, h in enumerate(headers2, 1):
            val = r.get(h, 0)
            # 变动面积列使用公式（当前期面积D列 - 对比期面积E列）
            if h == "变动面积":
                ws2.cell(row=row, column=i).value = f"=D{row}-E{row}"
                ws2.cell(row=row, column=i).number_format = '0.00'
            else:
                ws2.cell(row=row, column=i, value=val)
            ws2.cell(row=row, column=i).border = _THIN_BORDER
            # 统一对齐：使用状态名列左对齐，其余右对齐
            if h == "使用状态": 
                ws2.cell(row=row, column=i).alignment = _TEXT_ALIGN
            else:
                ws2.cell(row=row, column=i).alignment = _NUM_ALIGN
                ws2.cell(row=row, column=i).number_format = '0.00' if ("面积" in h or "额" in h) else '0'
            if h == "变动面积" and isinstance(val, (int, float)) and abs(val) > 0.01:
                ws2.cell(row=row, column=i).fill = _INCREASE_FILL if val > 0 else _DECREASE_FILL
        row += 1

    # 细分类汇总行
    detail_start = 4  # 细分类数据起始行
    detail_end = row - 1
    ws2.cell(row=row, column=1, value="合计").font = Font(bold=True, size=11)
    ws2.cell(row=row, column=1).border = _THIN_BORDER
    ws2.cell(row=row, column=1).alignment = _TEXT_ALIGN
    for i, h in enumerate(headers2, 1):
        if h == "使用状态": 
            continue
        col_letter = get_column_letter(i)
        ws2.cell(row=row, column=i).value = f"=SUM({col_letter}{detail_start}:{col_letter}{detail_end})"
        ws2.cell(row=row, column=i).border = _THIN_BORDER
        ws2.cell(row=row, column=i).font = Font(bold=True, size=11)
        ws2.cell(row=row, column=i).alignment = _NUM_ALIGN
        if "面积" in h or "额" in h:
            ws2.cell(row=row, column=i).number_format = '0.00'
        else:
            ws2.cell(row=row, column=i).number_format = '0'
    row += 2

    # 按大类汇总
    row += 2
    ws2.cell(row=row, column=1, value="按大类汇总").font = Font(bold=True, size=12)
    row += 1

    big_cat_headers = ["大类", "当前期面积", "对比期面积", "变动面积", "变动率(%)"]
    for i, h in enumerate(big_cat_headers, 1):
        ws2.cell(row=row, column=i, value=h)
    _style_header(ws2, row, len(big_cat_headers))
    row += 1

    for big_cat in ["可开发运作", "暂不可开发利用"]:
        area_c = stats_current.get(f"{big_cat}面积", 0)
        area_p = stats_previous.get(f"{big_cat}面积", 0)
        diff = round(area_c - area_p, 2)

        ws2.cell(row=row, column=1, value=big_cat).border = _THIN_BORDER
        ws2.cell(row=row, column=2, value=area_c).border = _THIN_BORDER
        ws2.cell(row=row, column=3, value=area_p).border = _THIN_BORDER
        # 变动面积 = 公式
        ws2.cell(row=row, column=4).value = f"=B{row}-C{row}"
        ws2.cell(row=row, column=4).number_format = '0.00'
        ws2.cell(row=row, column=4).border = _THIN_BORDER
        # 变动率 = 公式
        ws2.cell(row=row, column=5).value = f'=IF(C{row}=0,"-",D{row}/ABS(C{row}))'
        ws2.cell(row=row, column=5).number_format = '0.00%'
        ws2.cell(row=row, column=5).border = _THIN_BORDER
        if abs(diff) > 0.01:
            fill = _INCREASE_FILL if diff > 0 else _DECREASE_FILL
            ws2.cell(row=row, column=4).fill = fill
        row += 1

    _auto_width(ws2, len(headers2), row)

    # ============================================================
    # Sheet 3: 建筑类型变动汇总（与分类变动汇总格式一致）
    # ============================================================
    ws3 = wb.create_sheet("建筑类型变动汇总")

    ws3.merge_cells("A1:L1")
    ws3.cell(row=1, column=1, value="按建筑类型变动汇总").font = Font(bold=True, size=14)

    headers3 = [
        "建筑类型", "当前期宗数", "对比期宗数", "当前期面积", "对比期面积", "变动面积",
        "新增宗数", "新增面积", "删除宗数", "删除面积", "面积调整宗数", "面积调整额",
    ]
    for i, h in enumerate(headers3, 1):
        ws3.cell(row=3, column=i, value=h)
    _style_header(ws3, 3, len(headers3))

    # 复用 calc_building_type_summary
    bt_summary = calc_building_type_summary(df_current, df_previous)

    row = 4
    for item in bt_summary:
        ws3.cell(row=row, column=1, value=item["name"]).border = _THIN_BORDER
        ws3.cell(row=row, column=1).alignment = _TEXT_ALIGN
        for ci in range(2, 13):
            ws3.cell(row=row, column=ci).border = _THIN_BORDER
            ws3.cell(row=row, column=ci).alignment = _NUM_ALIGN
        ws3.cell(row=row, column=4).number_format = '0.00'
        ws3.cell(row=row, column=5).number_format = '0.00'
        ws3.cell(row=row, column=6).value = f"=D{row}-E{row}"
        ws3.cell(row=row, column=6).number_format = '0.00'
        ws3.cell(row=row, column=8).number_format = '0.00'
        ws3.cell(row=row, column=10).number_format = '0.00'
        ws3.cell(row=row, column=12).number_format = '0.00'
        ws3.cell(row=row, column=2, value=item["count_c"])
        ws3.cell(row=row, column=3, value=item["count_p"])
        ws3.cell(row=row, column=4, value=item["area_c"])
        ws3.cell(row=row, column=5, value=item["area_p"])
        ws3.cell(row=row, column=7, value=item["added_count"])
        ws3.cell(row=row, column=8, value=item["added_area"])
        ws3.cell(row=row, column=9, value=item["removed_count"])
        ws3.cell(row=row, column=10, value=item["removed_area"])
        ws3.cell(row=row, column=11, value=item["adj_count"])
        ws3.cell(row=row, column=12, value=item["adj_area"])

        if abs(item["area_diff"]) > 0.01:
            ws3.cell(row=row, column=6).fill = _INCREASE_FILL if item["area_diff"] > 0 else _DECREASE_FILL
        row += 1

    # 合计行
    detail_start = 4
    detail_end = row - 1
    ws3.cell(row=row, column=1, value="合计").font = Font(bold=True, size=11)
    ws3.cell(row=row, column=1).border = _THIN_BORDER
    ws3.cell(row=row, column=1).alignment = _TEXT_ALIGN
    for i, h in enumerate(headers3, 1):
        if h == "建筑类型":
            continue
        col_letter = get_column_letter(i)
        ws3.cell(row=row, column=i).value = f"=SUM({col_letter}{detail_start}:{col_letter}{detail_end})"
        ws3.cell(row=row, column=i).border = _THIN_BORDER
        ws3.cell(row=row, column=i).font = Font(bold=True, size=11)
        ws3.cell(row=row, column=i).alignment = _NUM_ALIGN
        if "面积" in h or "额" in h:
            ws3.cell(row=row, column=i).number_format = '0.00'
        else:
            ws3.cell(row=row, column=i).number_format = '0'

    _auto_width(ws3, len(headers3), row)

    # ============================================================
    # Sheet 4: 使用状态变动明细
    # ============================================================
    ws4 = wb.create_sheet("使用状态变动明细")
    ws4.cell(row=2, column=1, value=f"共 {len(changes_status_df)} 条变更记录")
    headers4 = [
        "资产编码", "资产分层（同系统名称）", "管理小组", "建筑类型",
        "变更类型", "建筑面积（当前期）", "建筑面积（对比期）", "建筑面积（变动）",
        "使用状态（当前期）", "使用状态（对比期）", "变更详情",
    ]
    _write_detail_sheet(ws4, headers4, changes_status_df, "使用状态变动明细（逐条）")

    # ============================================================
    # Sheet 5: 建筑类型变动明细
    # ============================================================
    ws5 = wb.create_sheet("建筑类型变动明细")
    ws5.cell(row=2, column=1, value=f"共 {len(changes_bt_df)} 条变更记录")
    headers5 = [
        "资产编码", "资产分层（同系统名称）", "管理小组", "使用状态",
        "变更类型", "建筑面积（当前期）", "建筑面积（对比期）", "建筑面积（变动）",
        "建筑类型（当前期）", "建筑类型（对比期）", "变更详情",
    ]
    _write_detail_sheet(ws5, headers5, changes_bt_df, "建筑类型变动明细（逐条）")

    # ============================================================
    # 保存
    # ============================================================
    # 处理文件已存在
    if Path(output_path).exists() and not force:
        base, ext = Path(output_path).stem, Path(output_path).suffix
        output_path = f"{base}_{time.strftime('%Y%m%d_%H%M%S')}{ext}"
        print(f"\n[提示] 输出文件已存在，自动另存为: {output_path}")

    # 处理文件被占用
    max_retries = 3
    for attempt in range(max_retries):
        try:
            wb.save(output_path)
            print(f"\n报告已保存: {output_path}")
            break
        except PermissionError:
            if attempt < max_retries - 1:
                print(f"\n[提示] 文件被占用，3秒后重试（{attempt + 1}/{max_retries}）...")
                time.sleep(3)
            else:
                base, ext = Path(output_path).stem, Path(output_path).suffix
                fallback = f"{base}_{time.strftime('%H%M%S')}{ext}"
                wb.save(fallback)
                print(f"\n[提示] 文件仍被占用，已另存为: {fallback}")


# ============================================================
# 七、主流程
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description="资产明细差异分析工具",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""使用示例:
  python asset_diff.py -c 当前期.xlsx -p 对比期.xlsx
  python asset_diff.py -c 当前期.xlsx -p 对比期.xlsx -o 报告.xlsx -f
  python asset_diff.py -c 当前期.xlsx -p 对比期.xlsx --scope 全部资产
""",
    )
    parser.add_argument("-c", "--current", required=True, help="当前期Excel文件路径")
    parser.add_argument("-p", "--previous", required=True, help="对比期Excel文件路径")
    parser.add_argument("-o", "--output", default=None,
                        help="输出文件路径（默认：[当前期文件名]_vs_[对比期文件名]_diff.xlsx）")
    parser.add_argument("-f", "--force", action="store_true",
                        help="覆盖已存在的输出文件（默认：自动加时间戳后缀）")
    parser.add_argument("--scope", default="自有资产", choices=["自有资产", "全部资产"],
                        help="统计范围（默认：自有资产）")
    args = parser.parse_args()

    print("=" * 60)
    print("资产明细差异分析")
    print("=" * 60)

    # 1. 读取数据
    print("\n[1/7] 读取数据...")
    df_current_raw, date_current = read_asset_file(args.current)
    df_previous_raw, date_previous = read_asset_file(args.previous)

    # 2. 数据检验
    print("\n[2/7] 数据检验...")
    validate_data(df_current_raw, df_previous_raw, date_current, date_previous)

    # 3. 数据预处理
    print("\n[3/7] 数据预处理...")
    for label, df in [("当前期", df_current_raw), ("对比期", df_previous_raw)]:
        map_status_category(df)
        standardize_building_type(df)

    # 4. 拆分自有/非自有
    print("\n[4/7] 拆分自有/非自有资产...")
    if args.scope == "自有资产":
        df_current, _ = split_self_owned(df_current_raw)
        df_previous, _ = split_self_owned(df_previous_raw)
        print(f"  当前期自有资产: {len(df_current)} 条")
        print(f"  对比期自有资产: {len(df_previous)} 条")
    else:
        df_current = df_current_raw
        df_previous = df_previous_raw

    # 5. 计算统计指标
    print("\n[5/7] 计算统计指标...")
    stats_current = calc_summary_stats(df_current, "当前期")
    stats_previous = calc_summary_stats(df_previous, "对比期")

    # 打印关键指标对比
    print(f"\n  {'指标':<20} {'当前期':>15} {'对比期':>15} {'变动':>15}")
    print(f"  {'-'*65}")
    key_indicators = [
        ("房产总面积", "㎡"), ("资产总宗数", "宗"),
        ("暂不可开发利用面积", "㎡"), ("可开发运作面积", "㎡"),
        ("已出租面积", "㎡"), ("空置面积", "㎡"),
        ("出租率", "%"), ("在约合同数", "份"),
    ]
    for key, unit in key_indicators:
        vc = stats_current.get(key, 0)
        vp = stats_previous.get(key, 0)
        diff = round(vc - vp, 2)
        print(f"  {key:<18} {vc:>14,.2f} {vp:>14,.2f} {diff:>+14,.2f} {unit}")

    # 6. 逐条比较
    print("\n[6/7] 逐条资产比较...")
    changes_status_df = compare_assets_by_status(df_current, df_previous)
    print(f"  使用状态变更记录: {len(changes_status_df)} 条")
    if len(changes_status_df) > 0:
        type_counts = changes_status_df["变更类型"].value_counts()
        for t, c in type_counts.items():
            print(f"    {t}: {c} 条")

    changes_bt_df = compare_assets_by_building_type(df_current, df_previous)
    print(f"  建筑类型变更记录: {len(changes_bt_df)} 条")
    if len(changes_bt_df) > 0:
        type_counts = changes_bt_df["变更类型"].value_counts()
        for t, c in type_counts.items():
            print(f"    {t}: {c} 条")

    # 7. 差异归因
    print("\n[7/7] 分类变动归因分析...")
    cat_changes_df = analyze_category_changes(df_current, df_previous)

    # 输出报告
    if args.output is None:
        name_c = Path(args.current).stem
        name_p = Path(args.previous).stem
        args.output = f"{name_c}_vs_{name_p}_diff.xlsx"
    write_excel_report(
        args.output,
        stats_current, stats_previous,
        changes_status_df, changes_bt_df, cat_changes_df,
        date_current, date_previous,
        df_current, df_previous,
        force=args.force,
    )

    print("\n完成！")


if __name__ == "__main__":
    main()
